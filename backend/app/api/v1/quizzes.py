import uuid
from fastapi import APIRouter, Depends, status, HTTPException, Response, Query, UploadFile, File, Form
import io
import pandas as pd
from openpyxl import load_workbook
from docx import Document
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import CurrentUser, get_current_user, get_db, PaginationParams, paginate, get_optional_user
from app.schemas.quiz import QuizCreate, QuizUpdate, QuizResponse, QuizListResponse, QuizDetailResponse
from app.schemas.question import QuestionCreate, QuestionResponse, ReorderQuestionsRequest
from app.repositories.quiz_repo import QuizRepository, QuestionRepository
from app.db.models.user import User

router = APIRouter(prefix="/quizzes", tags=["quizzes"])

@router.get("/public", response_model=QuizListResponse, summary="Lấy danh sách Quiz công khai")
async def get_public_quizzes(
    search: str | None = Query(None, description="Từ khóa tìm kiếm tiêu đề"),
    sort: str = Query("newest", description="Sắp xếp theo: newest, oldest"),
    pagination: PaginationParams = Depends(paginate),
    session: AsyncSession = Depends(get_db),
):
    """
    Truy vấn danh sách các bộ câu hỏi được cộng đồng chia sẻ công khai.
    """
    repo = QuizRepository(session)
    items, total = await repo.get_public(
        page=pagination.page, 
        size=pagination.size, 
        search=search, 
        sort=sort
    )
    return {
        "items": items,
        "total": total,
        "page": pagination.page,
        "size": pagination.size
    }

@router.post("", response_model=QuizResponse, status_code=status.HTTP_201_CREATED, summary="Tạo bộ câu hỏi mới")
async def create_quiz(
    data: QuizCreate,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    """
    Tạo một bộ câu hỏi mới cho người dùng hiện tại.
    """
    repo = QuizRepository(session)
    quiz = await repo.create(owner_id=current_user.id, data=data.model_dump())
    await session.commit()
    await session.refresh(quiz, ["questions"]) # To ensure question_count property works
    return quiz

@router.get("", response_model=QuizListResponse, summary="Lấy danh sách Quiz cá nhân")
async def get_quizzes(
    pagination: PaginationParams = Depends(paginate),
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    """
    Lấy toàn bộ danh sách Quiz mà người dùng hiện tại đã tạo.
    """
    repo = QuizRepository(session)
    items, total = await repo.get_by_owner(user_id=current_user.id, page=pagination.page, size=pagination.size)
    return {
        "items": items,
        "total": total,
        "page": pagination.page,
        "size": pagination.size
    }

@router.get("/{id}", response_model=QuizDetailResponse)
async def get_quiz(
    id: uuid.UUID,
    current_user: User | None = Depends(get_optional_user),
    session: AsyncSession = Depends(get_db),
):
    repo = QuizRepository(session)
    quiz = await repo.get_detail(id)
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    is_owner = current_user is not None and quiz.owner_id == current_user.id
    if not quiz.is_public and not is_owner:
        raise HTTPException(status_code=403, detail="Forbidden")
    
    response = QuizDetailResponse.model_validate(quiz)
    
    # Hide is_correct from non-owners
    if not is_owner:
        for q in response.questions:
            for opt in q.options:
                opt.is_correct = None
                
    return response

@router.patch("/{id}", response_model=QuizResponse)
async def update_quiz(
    id: uuid.UUID,
    data: QuizUpdate,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    repo = QuizRepository(session)
    quiz = await repo.get_by_id(id)
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
        
    if quiz.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    update_data = data.model_dump(exclude_unset=True)
    if update_data:
        quiz = await repo.update(id, update_data)
        await session.commit()
        await session.refresh(quiz, ["questions"]) # To ensure question_count property works
        
    return quiz

@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_quiz(
    id: uuid.UUID,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    repo = QuizRepository(session)
    quiz = await repo.get_by_id(id)
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
        
    if quiz.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    await repo.delete(id)
    await session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@router.post("/{id}/fork", response_model=QuizResponse)
async def fork_quiz(
    id: uuid.UUID,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    repo = QuizRepository(session)
    forked = await repo.fork(id, current_user.id)
    if not forked:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    await session.commit()
    await session.refresh(forked, ["questions"])
    return forked

@router.post("/{id}/questions", response_model=QuestionResponse, status_code=status.HTTP_201_CREATED)
async def create_question(
    id: uuid.UUID,
    data: QuestionCreate,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    quiz_repo = QuizRepository(session)
    quiz = await quiz_repo.get_by_id(id)
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
        
    if quiz.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    question_repo = QuestionRepository(session)
    data_dict = data.model_dump()
    data_dict["order_index"] = len(quiz.questions)
    
    question = await question_repo.create(quiz_id=id, data=data_dict)
    await session.commit()
    
    # Refresh to ensure relationships are loaded (though create already refreshes options)
    # The response_model needs options, which question_repo.create already eagerly loads.
    return question

@router.put("/{id}/questions/reorder", response_model=QuizDetailResponse)
async def reorder_questions(
    id: uuid.UUID,
    data: ReorderQuestionsRequest,
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    quiz_repo = QuizRepository(session)
    quiz = await quiz_repo.get_detail(id)
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
        
    if quiz.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    existing_ids = {q.id for q in quiz.questions}
    provided_ids = set(data.question_ids)
    
    if existing_ids != provided_ids or len(data.question_ids) != len(existing_ids):
        raise HTTPException(
            status_code=422, 
            detail="question_ids must contain exactly all current question IDs without duplicates"
        )
        
    question_repo = QuestionRepository(session)
    await question_repo.reorder(id, data.question_ids)
    await session.commit()
    
    # Fetch again to get updated order and relations
    await session.refresh(quiz, ["questions"])
    return QuizDetailResponse.model_validate(quiz)

@router.post("/import-excel", response_model=QuizResponse, status_code=status.HTTP_201_CREATED)
async def import_quiz_from_excel(
    title: str = Form(...),
    description: str | None = Form(None),
    is_public: bool = Form(False),
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        
        questions_to_create = []
        
        # Duyệt từ dòng thứ 2 (dòng 1 là header)
        for row in sheet.iter_rows(min_row=2, values_only=False):
            # Cột 1 (index 0): Câu hỏi
            q_text = row[0].value
            if not q_text:
                continue
                
            options = []
            # Cột 2-5 (index 1-4): Các lựa chọn
            for i in range(1, 5):
                opt_text = row[i].value
                if opt_text is None:
                    continue
                
                # Kiểm tra màu nền ô. Nếu có màu (không phải trắng/không màu) thì là đáp án đúng
                # Lưu ý: PatternFill.fgColor có thể là '00000000' (trong suốt)
                is_correct = False
                fill = row[i].fill
                if fill and fill.start_color and fill.start_color.index != '00000000' and fill.start_color.rgb != '00000000':
                    is_correct = True
                
                options.append({
                    "option_text": str(opt_text),
                    "is_correct": is_correct
                })
            
            # Cột 6 (index 5): Thời gian (giây)
            time_limit = row[5].value if len(row) > 5 and row[5].value else 30
            # Cột 7 (index 6): Điểm
            points = row[6].value if len(row) > 6 and row[6].value else 1000
            
            questions_to_create.append({
                "question_text": str(q_text),
                "question_type": "multiple_choice",
                "time_limit_secs": int(time_limit),
                "points": int(points),
                "options": options
            })

        if not questions_to_create:
            raise HTTPException(status_code=400, detail="No valid questions found in Excel file")

        # Tạo Quiz
        quiz_repo = QuizRepository(session)
        quiz = await quiz_repo.create(owner_id=current_user.id, data={
            "title": title,
            "description": description,
            "is_public": is_public
        })
        
        # Tạo Câu hỏi
        question_repo = QuestionRepository(session)
        for idx, q_data in enumerate(questions_to_create):
            q_data["order_index"] = idx
            await question_repo.create(quiz_id=quiz.id, data=q_data)
            
        await session.commit()
        await session.refresh(quiz, ["questions"])
        return quiz
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@router.post("/import-word", response_model=QuizResponse, status_code=status.HTTP_201_CREATED)
async def import_quiz_from_word(
    title: str = Form(...),
    description: str | None = Form(None),
    is_public: bool = Form(False),
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith('.docx'):
        raise HTTPException(status_code=400, detail="Only .docx files are supported")

    try:
        content = await file.read()
        doc = Document(io.BytesIO(content))
        
        questions_to_create = []
        paragraphs = [p for p in doc.paragraphs if p.text.strip()]
        
        # Giả định cấu trúc: 1 câu hỏi + 4 đáp án = 5 paragraph
        i = 0
        while i < len(paragraphs):
            q_text = paragraphs[i].text.strip()
            options = []
            
            # Lấy 4 đoạn tiếp theo làm đáp án
            for j in range(1, 5):
                if i + j < len(paragraphs):
                    opt_para = paragraphs[i+j]
                    opt_text = opt_para.text.strip()
                    
                    # Kiểm tra highlight trong các run của paragraph
                    is_correct = False
                    for run in opt_para.runs:
                        if run.font.highlight_color is not None:
                            is_correct = True
                            break
                    
                    # Nếu text có dạng "A. Nội dung", "B. Nội dung", ta có thể strip tiền tố nếu muốn
                    # Ở đây mình giữ nguyên hoặc xử lý nhẹ
                    clean_opt_text = opt_text
                    if len(opt_text) > 2 and opt_text[1:3] == ". ":
                        clean_opt_text = opt_text[3:]

                    options.append({
                        "option_text": clean_opt_text,
                        "is_correct": is_correct
                    })
            
            if q_text and options:
                questions_to_create.append({
                    "question_text": q_text,
                    "question_type": "multiple_choice",
                    "time_limit_secs": 30,
                    "points": 1000,
                    "options": options
                })
            
            i += 5 # Nhảy sang cụm tiếp theo

        if not questions_to_create:
            raise HTTPException(status_code=400, detail="No valid questions found in Word file. Ensure 1 question followed by 4 options.")

        # Tạo Quiz & Questions (Reuse logic)
        quiz_repo = QuizRepository(session)
        quiz = await quiz_repo.create(owner_id=current_user.id, data={
            "title": title,
            "description": description,
            "is_public": is_public
        })
        
        question_repo = QuestionRepository(session)
        for idx, q_data in enumerate(questions_to_create):
            q_data["order_index"] = idx
            await question_repo.create(quiz_id=quiz.id, data=q_data)
            
        await session.commit()
        await session.refresh(quiz, ["questions"])
        return quiz

    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing Word file: {str(e)}")
